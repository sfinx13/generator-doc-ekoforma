from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def create_table(ws, start_row, start_col, title, data):
    # Calculer les cellules de début et de fin pour la fusion
    start_cell = ws.cell(row=start_row, column=start_col).coordinate
    end_cell = ws.cell(row=start_row, column=start_col + 6).coordinate
    ws.merge_cells(f'{start_cell}:{end_cell}')
    
    # Ajouter le titre
    ws[start_cell] = title

    # Appliquer les styles : police Calibri, taille 16, en gras, centré
    font = Font(name='Calibri', size=16, bold=True)
    alignment = Alignment(horizontal='center', vertical='center')
    
    ws[start_cell].font = font
    ws[start_cell].alignment = alignment

    ws.row_dimensions[start_row].height = 30  # Ajustez cette valeur selon vos besoins
    
    # Définir une bordure de 2pt
    border_style = Side(border_style='medium', color='000000')
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    # Appliquer la bordure à toutes les cellules fusionnées
    for col in range(start_col, start_col + 7):
        cell = ws.cell(row=start_row, column=col)
        cell.border = border

    participants_section = False

    # Ajouter les données
    for row_index, row_data in enumerate(data, start=start_row + 1):
        for col_index, cell_value in enumerate(row_data, start=start_col):
            cell = ws.cell(row=row_index, column=col_index)
            cell.value = cell_value

            # Détecter la section des participants
            if cell_value == "PARTICIPANTS":
                participants_section = True

            # Appliquer les styles spécifiques
            if row_index == start_row + 1 or row_index == start_row + 2:
                cell.font = Font(name='Calibri', size=14)
                if col_index == start_col + 1:
                    cell.font = Font(name='Calibri', size=14, bold=True)
                ws.row_dimensions[row_index].height = 20  # Augmenter légèrement la hauteur de la ligne

            if row_index == start_row + 4:
                cell.font = Font(name='Calibri', size=12, bold=True)
                cell.fill = PatternFill(start_color='DEEBF7', end_color='DEEBF7', fill_type='solid')
            
            if row_index == start_row + 5:
                cell.font = Font(name='Calibri', size=12, bold=True)

            if row_index == start_row + 7:
                cell.font = Font(name='Calibri', size=12, bold=True)
                cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[row_index].height = 25
                for col in range(start_col, start_col + 7):
                    cell = ws.cell(row=row_index, column=col)
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
            
            if row_index == start_row + 9:
                if col_index == start_col:
                    cell.font = Font(name='Calibri', size=12, bold=True)

            if row_index == start_row + 10 or row_index == start_row + 14:
                cell.font = Font(name='Calibri', size=12, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            if row_index == start_row + 11 or row_index == start_row + 15:
                cell.font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)  # Police de couleur blanche
                cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[row_index].height = 30  # Augmenter la hauteur de la ligne

                if col_index < start_col + 4:
                    cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')  # Couleur bleue
                else:
                    cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')  # Couleur jaune

            if row_index == start_row + 12 and col_index in [start_col, start_col + 1]:
                cell.alignment = Alignment(horizontal='left')
            
            if row_index == start_row + 12 and col_index in [start_col + 4, start_col + 5, start_col + 6]:
                cell.alignment = Alignment(horizontal='right')

            if row_index == start_row + 11:
                # Bordure haute pour ligne 12
                cell.border = Border(top=border_style)
                if col_index == start_col:
                    cell.border = Border(top=border_style, left=border_style)
                if col_index == start_col + 6:
                    cell.border = Border(top=border_style, right=border_style)
            elif row_index == start_row + 12:
                # Bordure basse pour ligne 13
                cell.border = Border(bottom=border_style)
                if col_index == start_col:
                    cell.border = Border(bottom=border_style, left=border_style)
                if col_index == start_col + 6:
                    cell.border = Border(bottom=border_style, right=border_style)
            
            # Colorer les lignes vides après la 17ème ligne
            if participants_section and not any(row_data):
                if col_index < start_col + 4:
                    cell.fill = PatternFill(start_color='DEEBF7', end_color='DEEBF7', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')

            # Appliquer la police Trebuchet MS aux participants
            if participants_section and any(row_data):
                cell.font = Font(name='Trebuchet MS', size=12)
                if col_index <= start_col + 1:
                    cell.alignment = Alignment(horizontal='left')
                elif col_index == start_col + 2 or col_index >= start_col + 4:
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='center')


            # Ajouter bordure gauche pour la première cellule de chaque ligne de participant
            if participants_section and col_index == start_col and cell_value != "PARTICIPANTS":
                cell.border = Border(left=border_style)

            # Ajouter bordure droite pour la dernière cellule de chaque ligne de participant
            if participants_section and col_index == start_col + 6 and cell_value != "PARTICIPANTS":
                cell.border = Border(right=border_style)


            # Stop styling after the "Je soussigné(e)" line
            if cell_value and "Je sousign" in cell_value:
                merge_start_cell = ws.cell(row=row_index, column=start_col).coordinate
                merge_end_cell = ws.cell(row=row_index, column=start_col + 6).coordinate
                ws.merge_cells(f'{merge_start_cell}:{merge_end_cell}')
                ws[merge_start_cell].border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)
                ws[merge_start_cell].alignment = Alignment(vertical="center")
                ws[merge_start_cell].font = Font(name='Calibri', size=12)
                ws.row_dimensions[row_index].height = 60  # Augmenter légèrement la hauteur de la ligne
                participants_section = False

            if cell_value and "Article 441-1 du code pénal:" in cell_value:
                
                merge_start_cell = ws.cell(row=row_index, column=start_col).coordinate
                merge_end_cell = ws.cell(row=row_index, column=start_col + 6).coordinate
                ws.merge_cells(f'{merge_start_cell}:{merge_end_cell}')
                ws[merge_start_cell].border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)
                ws[merge_start_cell].alignment = Alignment(vertical="center")
                ws[merge_start_cell].font = Font(name='Calibri', size=11)
                ws.row_dimensions[row_index].height = 40  # Augmenter légèrement la hauteur de la ligne
                bold_text = "Article 441-1 du code pénal:"
                normal_text = cell_value.replace(bold_text, "")    
                # ws.write(row_index, col_index, bold_text, style_bold)    
                # ws.write_merge(row_index, row_index, col_index, col_index, bold_text + normal_text, style_normal)
    
    # Définir les largeurs des colonnes
    ws.column_dimensions[ws.cell(row=start_row + 1, column=start_col).column_letter].width = 35
    ws.column_dimensions[ws.cell(row=start_row + 1, column=start_col + 1).column_letter].width = 35
    ws.column_dimensions[ws.cell(row=start_row + 1, column=start_col + 2).column_letter].width = 20
    ws.column_dimensions[ws.cell(row=start_row + 1, column=start_col + 3).column_letter].width = 15
    ws.column_dimensions[ws.cell(row=start_row + 1, column=start_col + 4).column_letter].width = 30
    ws.column_dimensions[ws.cell(row=start_row + 1, column=start_col + 5).column_letter].width = 30
    ws.column_dimensions[ws.cell(row=start_row + 1, column=start_col + 6).column_letter].width = 38

    # Fusionner les colonnes spécifiées pour la cinquième ligne
    ws.merge_cells(start_row=start_row + 4, start_column=start_col + 2, end_row=start_row + 4, end_column=start_col + 3)
    # Fusionner les colonnes spécifiées pour la huitième ligne
    ws.merge_cells(start_row=start_row + 7, end_row=start_row + 7, start_column=start_col, end_column=start_col + 6)
    # Fusionner les colonnes spécifiées pour la onzieme ligne
    ws.merge_cells(start_row=start_row + 10, end_row=start_row + 10, start_column=start_col, end_column=start_col + 6)
    ws.merge_cells(start_row=start_row + 11, start_column=start_col + 2, end_row=start_row + 11, end_column=start_col + 3)
    ws.merge_cells(start_row=start_row + 12, start_column=start_col + 2, end_row=start_row + 12, end_column=start_col + 3)

    ws.merge_cells(start_row=start_row + 14, end_row=start_row + 14, start_column=start_col, end_column=start_col + 6)


# Fonction pour créer les données pour chaque demi-journée
def create_data_for_meeting(meeting, participants):
    formateur_morning = participants['formateur']['morning']
    data = [
        ["Nom de l'organisme :", "EKOFORMA"],
        ["Identifiant :", "99LH"],
        [],
        [meeting[1], "", meeting[2], "", meeting[3], "Volume horaire déclaré :", meeting[4]],
        [meeting[5]],  # Titre de la formation
        [],
        ["Adresse url / logiciel utilisé : ZOOM"],
        [],
        [meeting[6], meeting[7], meeting[8], "", "1ère demi-journée"],
        ["INTERVENANTS"],
        ["NOM", "Prénom", "N°RPPS ou Adeli", "", "Heure de la 1ère connexion", "Heure de la dernière \nconnexion", "Total du temps connecté réalisé (en \n minutes)"],
        [formateur_morning[0].split()[1], formateur_morning[0].split()[0], "", "", meeting[9], meeting[10], str(meeting[12])],
        [],
        ["PARTICIPANTS"],
        ["NOM", "Prénom", "N°RPPS ou Adeli", "Financeur", "Heure de la 1ère connexion", "Heure de la dernière \nconnexion", "Total du temps connecté réalisé (en \n minutes)"],
    ]

    for email, participant in participants.items():
        if email != 'formateur':  # Ignorer le formateur dans les participants
            if "MATIN" in meeting[4]:
                morning_session = participant['morning']
                data.append([morning_session[0], morning_session[1], "", "ANDPC", morning_session[3], morning_session[4], str(morning_session[5])])
                data.append(["", "", "", "", "", "", ""])  # Ligne vide entre chaque participant
            else:
                afternoon_session = participant['afternoon']
                data.append([afternoon_session[0], afternoon_session[1], "", "ANDPC", afternoon_session[3], afternoon_session[4], str(afternoon_session[5])])
                data.append(["", "", "", "", "", "", ""])  # Ligne vide entre chaque participant


     # Ajout des lignes finales
    data.append([
        "Je sousignée(é) ZAFER MOHAMED agissant en ma qualité de Président, Directeur Général de l'organisme EKOFORMA atteste que les personnes dont les noms figurent ci-dessus ont suivi les séquences de la classe virtuelle de l'action ou du programme dont le numéro et la session sont indiqués en haut à gauche de cette attestation. Je joins en complément de cette attestation l'ensemble des logs informatiques issus de ma plateforme."
    ])
    data.append(["Cachet de l'organisme : ", "", "", "", "", "", ""]) # Ne pas oublié la date
    data.append([])
    data.append([])
    data.append([])
    data.append([
        "Article 441-1 du code pénal: \"Constitue un faux toute altération frauduleuse de la vérité, de nature à causer un préjudice et accomplie par quelque moyen que ce soit, dans un écrit ou tout autre support d'expression de la pensée qui a pour objet ou qui peut avoir pour effet d'établir la preuve d'un droit ou d'un fait ayant des conséquences juridiques. Le faux et l'usage de faux sont punis de trois ans d'emprisonnement et de 45 000 euros d'amende.\""
    ])

    return data


# Fonction pour générer les tableaux dans la feuille Excel
def generate_tables_for_each_meeting(filename, data_structure):
    start_row = 1
    start_col = 1
    wb = Workbook()
    ws = wb.active

    for day_key in ['date_debut', 'date_fin']:
        day_info = data_structure.get(day_key)
        if day_info:
            for meeting in day_info['meetings']:
                participants = day_info['participants']  # Récupérer les participants pour cette journée
                data = create_data_for_meeting(meeting, participants)
                
                # Appel de la méthode create_table avec le start_row actuel
                create_table(ws, start_row=start_row, start_col=start_col, title="Synthèse de suivi de classe virtuelle", data=data)
                
                # Calcul du nombre de lignes dans le tableau actuel
                num_lines = len(data)
                
                # Incrémentation de start_row pour le prochain tableau
                start_row += num_lines + 10  # Ajout de 10 lignes d'espace entre chaque tableau
    
    wb.save("./public/synthese_de_suivi_classe_virtuelle_{}".format(filename))
