from openpyxl import load_workbook
from datetime import datetime, timedelta

def parse_sheet(filepath, sheet_name):
    data_sheet = []
    try:
        wb = load_workbook(filename=filepath)
        
        sheet = wb[sheet_name]

        def extract_data(sheet):
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            return data

        data_sheet = extract_data(sheet)
    except:
        print("Erreur : Le fichier n'est pas supporté ou est corrompu.")

    return data_sheet

def create_formation(filepath):
    formation = {}
    try:
        data = parse_sheet(filepath, 'formation')
        for row in data:
            formation[row[0].lower()] = row[1]
    
        days, month, year = formation.get('date').split('/')
        day_start = days
        day_end = None

        if ("-" in days):
            day_start, day_end = days.split('-')
        
        formation['date_debut'] = datetime.strptime(f"{day_start}/{month}/{year}", "%d/%m/%y")
        
        if (day_end is not None):
            formation['date_fin'] = datetime.strptime(f"{day_end}/{month}/{year}", "%d/%m/%y")
        
        date_debut = formation['date_debut']
        date_fin = formation.get('date_fin', date_debut) 
        volumes_horaires = []
        current_date = date_debut
        while current_date <= date_fin:
            date_str = current_date.strftime('%d/%m/%Y')
            volumes_horaires.append(f"{date_str} : 3 H MATIN")
            volumes_horaires.append(f"{date_str} : 4 H APRÈS-MIDI")
            current_date += timedelta(days=1)
            
        formation['volumes_horaires'] = volumes_horaires
    
    except Exception as e:
        print(f"Erreur: {e}")

    return formation

def create_participants(filepath):
    participants = []
    try:
        data = parse_sheet(filepath, 'participants')        
        for row in data:
            participant = {}
            participant['civilite'] = row[0]
            participant['nom_complet'] = "{} {}".format(row[1], row[2])
            participant['nom'] = row[1]
            participant['prenoms'] = row[2]
            participant['prenom'] = row[2].split()[0]
            participant['email'] = row[3]
            participant['rpps'] = row[4]
            participant['phone'] = row[5]
            participant['financement'] = row[6]
            participants.append(participant)
    except Exception as e:
        print(f"Erreur: {e}")

    return participants