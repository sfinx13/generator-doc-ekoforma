from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta
import random
import string

def generate_random_string(length=12):
    return '9' + ''.join(random.choices(string.digits, k=length-1))

def generate_random_time(workshop_date, start_hour, start_minute, end_hour, end_minute):
    start_time = timedelta(hours=start_hour, minutes=start_minute)
    end_time = timedelta(hours=end_hour, minutes=end_minute)
    random_time = start_time + (end_time - start_time) * random.random()
    
    return datetime.combine(workshop_date, (datetime.min + random_time).time())

def calculate_duration(start_time, end_time):
    return int((end_time - start_time).total_seconds() / 60)

def create_zoom_timesheet(filepath, formation, participants):
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"

    def generate_meetings_and_participants(date_formation, row_header_title = 1):
        workshop_number = generate_random_string()
        merged_text = f"participants_{workshop_number}_zoom"

        ws.merge_cells(start_row=row_header_title, start_column=1, end_row=row_header_title, end_column=7)
        merged_cell = ws.cell(row=row_header_title, column=1)
        merged_cell.value = merged_text
        merged_cell.font = Font(name='Calibri', size=11, bold=True)
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        merged_cell.border = thin_border

        headers = [
            "N° de réunion", 
            "Sujet", 
            "Heure de début", 
            "Heure de fin", 
            "Adresse e-mail de l'utilisateur", 
            "Durée (minutes)", 
            "Participants"
        ]

        ws.append(headers)

        header_font = Font(name='Calibri', size=11, bold=True, color='000000') 
        header_fill = PatternFill(start_color='BEC0BF', end_color='BEC0BF', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=1, max_col=7, max_row=ws.max_row):
            if row[0].value == "N° de réunion":
                for cell in row:
                    cell.font = header_font 
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

        column_widths = [35, 60, 20, 20, 30, 15, 15]
        for i, column_width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = column_width

        meetings = []
        num_participants = len(participants)
        
        start_time_morning = generate_random_time(date_formation, 8, 45, 8, 58)
        end_time_morning = generate_random_time(date_formation, 12, 10, 12, 15)
        meetings.append([
            workshop_number,
            "Formation - {}".format(formation['titre']),
            start_time_morning.strftime("%d/%m/%y %H:%M:%S"),
            end_time_morning.strftime("%d/%m/%y %H:%M:%S"),
            "classe{}@ekoforma.com".format(formation['classe']).lower(),
            calculate_duration(start_time_morning, end_time_morning),
            num_participants
        ])
        
        start_time_afternoon = generate_random_time(date_formation, 13, 15, 13, 28)
        end_time_afternoon = generate_random_time(date_formation,17, 35, 17, 45)
        meetings.append([
            workshop_number,
            "Formation - {}".format(formation['titre']),
            start_time_afternoon.strftime("%d/%m/%y %H:%M:%S"),
            end_time_afternoon.strftime("%d/%m/%y %H:%M:%S"),
            "classe{}@ekoforma.com".format(formation['classe']).lower(),
            calculate_duration(start_time_afternoon, end_time_afternoon),
            num_participants
        ])

        for meeting in meetings:
            ws.append(meeting)
            ws.row_dimensions[ws.max_row].height = 40


        participants_headers = [
            "Nom (nom original)", 
            "Adresse e-mail de l’utilisateur", 
            "Heure d’arrivée", 
            "Heure de départ", 
            "Durée (minutes)", 
            "Invité", 
            "Salle d’attente"
        ]

        ws.append(['empty'])
        ws.append(participants_headers)
        ws.row_dimensions[ws.max_row].height = 25

        ws.append([
            'Formateur Ekoforma',
            "classe{}@ekoforma.com".format(formation['classe']).lower(),
            start_time_morning.strftime("%d/%m/%y %H:%M:%S"),
            end_time_morning.strftime("%d/%m/%y %H:%M:%S"),
            calculate_duration(start_time_morning, end_time_morning),
            'Non',
            'Non'
        ])
        ws.row_dimensions[ws.max_row].height = 25

        ws.append([
            'Formateur Ekoforma',
            "classe{}@ekoforma.com".format(formation['classe']).lower(),
            start_time_afternoon.strftime("%d/%m/%y %H:%M:%S"),
            end_time_afternoon.strftime("%d/%m/%y %H:%M:%S"),
            calculate_duration(start_time_afternoon, end_time_afternoon),
            'Non',
            'Non'
        ])
        ws.row_dimensions[ws.max_row].height = 25

        start_time_morning_participant = generate_random_time(date_formation, 9, 0, 9, 7)
        end_time_morning_participant = end_time_morning
        start_time_afternoon_participant = generate_random_time(date_formation, 13, 30, 13, 37)
        end_time_afternoon_participant = end_time_afternoon
        ws.append(['empty'])
        for participant in participants:
            for row in range(2):
                if (row == 0):
                    ws.append([
                        "{}".format(participant['nom_complet']),
                        participant['email'],
                        start_time_morning_participant.strftime("%d/%m/%y %H:%M:%S"),
                        end_time_morning_participant.strftime("%d/%m/%y %H:%M:%S"),
                        calculate_duration(start_time_morning_participant, end_time_morning_participant),
                        "Oui",
                        "Oui"
                    ])
                if (row == 1):
                    ws.append([
                        "{}".format(participant['nom_complet']),
                        participant['email'],
                        start_time_afternoon_participant.strftime("%d/%m/%y %H:%M:%S"),
                        end_time_afternoon_participant.strftime("%d/%m/%y %H:%M:%S"),
                        calculate_duration(start_time_afternoon_participant, end_time_afternoon_participant),
                        "Oui",
                        "Oui"
                    ])
                ws.row_dimensions[ws.max_row].height = 25

                    
        
    generate_meetings_and_participants(formation['date_debut'])
    
    ws.append([])
    ws.row_dimensions[ws.max_row + 1].height = 80
    ws.merge_cells(start_row=ws.max_row + 1, start_column=1, end_row=ws.max_row + 1, end_column=7)

    if (formation.get('date_fin')):
        generate_meetings_and_participants(formation['date_fin'], ws.max_row + 1)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 'empty' or cell.value:
                cell.border = thin_border
    
    wb.save("zoom_timesheet_{}".format(filepath))

    return wb, ws


