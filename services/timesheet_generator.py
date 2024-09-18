from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.page import PageMargins
from datetime import datetime, timedelta
import random
import string
import os

def setup_excel_for_pdf(ws):
    """
    Configure la mise en page de la feuille Excel pour qu'elle s'ajuste à une page PDF.
    """
    # Définir les marges de la page
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    # Configurer la mise à l'échelle pour ajuster à une seule page
    ws.print_options.horizontalCentered = True  # Centrer horizontalement
    ws.print_options.verticalCentered = True    # Centrer verticalement
    
    # Ajuster à une seule page
    ws.page_setup.fitToWidth = 1  # Réduire pour tenir sur la largeur d'une page
    ws.page_setup.fitToHeight = 1  # Réduire pour tenir sur la hauteur d'une page

def generate_random_string(length=12):
    return '9' + ''.join(random.choices(string.digits, k=length-1))

def generate_random_time(workshop_date, start_hour, start_minute, end_hour, end_minute):
    start_time = timedelta(hours=start_hour, minutes=start_minute)
    end_time = timedelta(hours=end_hour, minutes=end_minute)
    random_time = start_time + (end_time - start_time) * random.random()
    
    return datetime.combine(workshop_date, (datetime.min + random_time).time())

def calculate_duration(start_time, end_time):
    return int((end_time - start_time).total_seconds() / 60)

def create_zoom_timesheet(filepath, formation, participants):
    wb = Workbook()
    full_meetings_and_participants = {}
    workshop_number = generate_random_string()
    merged_text = f"participants_{workshop_number}_zoom"

    def generate_meetings_and_participants(date_formation, ws: Worksheet):
        meetings_and_participants = {}
        virtual_meetings_and_participants = {}

        # Fusion des cellules pour la première ligne
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        merged_cell = ws.cell(row=1, column=1)
        merged_cell.value = merged_text
        merged_cell.font = Font(name='Helvetica Neue', size=11, bold=True)
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        merged_cell.border = thin_border

        headers = [
            "N° de réunion", 
            "Sujet", 
            "Heure de début", 
            "Heure de fin", 
            "Adresse e-mail de l'utilisateur", 
            "Durée (minutes)", 
            "Participants"
        ]

        ws.append(headers)

        header_font = Font(name='Helvetica Neue', size=11, bold=True, color='000000') 
        header_fill = PatternFill(start_color='BEC0BF', end_color='BEC0BF', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=1, max_col=7, max_row=ws.max_row):
            if row[0].value == "N° de réunion":
                for cell in row:
                    cell.font = header_font 
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

        column_widths = [25, 40, 20, 20, 30, 20, 15]
        for i, column_width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = column_width

        meetings = []
        virtual_meetings = []
        num_participants = len(participants)
        
        start_time_morning = generate_random_time(date_formation, 8, 45, 8, 58)
        end_time_morning = generate_random_time(date_formation, 12, 10, 12, 15)
        virtual_meetings.append([
            workshop_number,
            "N° Action / Programme : {}".format(formation['session']),
            "N° de session : {}".format(formation.get('code')),
            "N° de l’unité: 1",
            "{} : 3 H MATIN".format(start_time_morning.strftime('%d/%m/%Y')),
            formation['titre'],
            "Date de la vacation : {}".format(start_time_morning.strftime("%d/%m/%y")),
            "Heure de début : {}".format(start_time_morning.strftime("%H:%M")),
            "Heure de Fin : {}".format(end_time_morning.strftime("%H:%M")),
            start_time_morning.strftime("%d/%m/%y %H:%M:%S"),
            end_time_morning.strftime("%d/%m/%y %H:%M:%S"),
            "classe{}@ekoforma.com".format(formation['classe']).lower(),
            calculate_duration(start_time_morning, end_time_morning),
            num_participants + 1 
        ])

        meetings.append([
            workshop_number,
            "Formation - {}".format(formation['titre']),
            start_time_morning.strftime("%d/%m/%y %H:%M:%S"),
            end_time_morning.strftime("%d/%m/%y %H:%M:%S"),
            "classe{}@ekoforma.com".format(formation['classe']).lower(),
            calculate_duration(start_time_morning, end_time_morning),
            num_participants + 1
        ])
        
        start_time_afternoon = generate_random_time(date_formation, 13, 15, 13, 28)
        end_time_afternoon = generate_random_time(date_formation,17, 35, 17, 45)
        virtual_meetings.append([
            workshop_number,
            "N° Action / Programme : {}".format(formation['session']),
            "N° de session : {}".format(formation.get('code')),
            "N° de l’unité: 1",
            "{} : 4 H APRÈS-MIDI".format(start_time_afternoon.strftime('%d/%m/%Y')),
            formation['titre'],
            "Date de la vacation : {}".format(start_time_afternoon.strftime("%d/%m/%y")),
            "Heure de début : {}".format(start_time_afternoon.strftime("%H:%M")),
            "Heure de Fin : {}".format(end_time_afternoon.strftime("%H:%M")),
            start_time_afternoon.strftime("%d/%m/%y %H:%M:%S"),
            end_time_afternoon.strftime("%d/%m/%y %H:%M:%S"),
            "classe{}@ekoforma.com".format(formation['classe']).lower(),
            calculate_duration(start_time_afternoon, end_time_afternoon),
            num_participants + 1
        ])

        meetings.append([
            workshop_number,
            "Formation - {}".format(formation['titre']),
            start_time_afternoon.strftime("%d/%m/%y %H:%M:%S"),
            end_time_afternoon.strftime("%d/%m/%y %H:%M:%S"),
            "classe{}@ekoforma.com".format(formation['classe']).lower(),
            calculate_duration(start_time_afternoon, end_time_afternoon),
            num_participants + 1
        ])


        meetings_and_participants['meetings'] = meetings
        virtual_meetings_and_participants['meetings'] = meetings
        for meeting in meetings:
            ws.append(meeting)
            ws.row_dimensions[ws.max_row].height = 60
            workshop_cell = ws.cell(row=ws.max_row, column=1)
            workshop_cell.alignment = Alignment(horizontal='right')

            subject_cell = ws.cell(row=ws.max_row, column=2)
            subject_cell.alignment = Alignment(wrap_text=True)


        participants_headers = [
            "Nom (nom original)", 
            "Adresse e-mail de l’utilisateur", 
            "Heure d’arrivée", 
            "Heure de départ", 
            "Durée (minutes)", 
            "Invité", 
            "Salle d’attente"
        ]

        ws.append(['empty', '', '', '', '', '', 'empty'])
        ws.append(participants_headers)
        ws.row_dimensions[ws.max_row].height = 25

        ws.append([
            'Formateur Ekoforma',
            "classe{}@ekoforma.com".format(formation['classe']).lower(),
            start_time_morning.strftime("%d/%m/%y %H:%M:%S"),
            end_time_morning.strftime("%d/%m/%y %H:%M:%S"),
            calculate_duration(start_time_morning, end_time_morning),
            'Non',
            'Non'
        ])

        if 'participants' not in meetings_and_participants:
            meetings_and_participants['participants'] = {}
            virtual_meetings_and_participants['participants'] = {}
            if 'formateur' not in meetings_and_participants['participants']:
                meetings_and_participants['participants']['formateur'] = {}
                virtual_meetings_and_participants['participants']['formateur'] = {}

        meetings_and_participants['participants']['formateur']['morning'] = [
            formation['formateur'],
            formation['session'],
            "classe{}@ekoforma.com".format(formation['classe']).lower(),
            start_time_morning.strftime("%d/%m/%y %H:%M:%S"),
            end_time_morning.strftime("%d/%m/%y %H:%M:%S"),
            calculate_duration(start_time_morning, end_time_morning)
        ]

        ws.row_dimensions[ws.max_row].height = 25

        ws.append([
            'Formateur Ekoforma',
            "classe{}@ekoforma.com".format(formation['classe']).lower(),
            start_time_afternoon.strftime("%d/%m/%y %H:%M:%S"),
            end_time_afternoon.strftime("%d/%m/%y %H:%M:%S"),
            calculate_duration(start_time_afternoon, end_time_afternoon),
            'Non',
            'Non'
        ])

        meetings_and_participants['participants']['formateur']['afternoon'] = [
            formation['formateur'],
            formation['session'],
            "classe{}@ekoforma.com".format(formation['classe']).lower(),
            start_time_afternoon.strftime("%d/%m/%y %H:%M:%S"),
            end_time_afternoon.strftime("%d/%m/%y %H:%M:%S"),
            calculate_duration(start_time_afternoon, end_time_afternoon),
        ]

        ws.row_dimensions[ws.max_row].height = 25

        start_time_morning_participant = generate_random_time(date_formation, 9, 0, 9, 7)
        end_time_morning_participant = end_time_morning
        start_time_afternoon_participant = generate_random_time(date_formation, 13, 30, 13, 37)
        end_time_afternoon_participant = end_time_afternoon
        ws.append(['empty', '', '', '', '', '', 'empty'])
        for participant in participants:
            for row in range(2):
                if (row == 0):
                    ws.append([
                        "{} {}".format(participant['nom'].lower(), participant['prenom'].lower()),
                        participant['email'],
                        start_time_morning_participant.strftime("%d/%m/%y %H:%M:%S"),
                        end_time_morning_participant.strftime("%d/%m/%y %H:%M:%S"),
                        calculate_duration(start_time_morning_participant, end_time_morning_participant),
                        "Oui",
                        "Oui"
                    ])
                    if participant['email'] not in meetings_and_participants['participants']:
                        meetings_and_participants['participants'][participant['email']] = {}
                    
                    meetings_and_participants['participants'][participant['email']]['morning'] = [
                        "{}".format(participant['nom_complet']),
                        participant['prenom'],
                        participant['email'],
                        start_time_morning_participant.strftime("%d/%m/%y %H:%M:%S"),
                        end_time_morning_participant.strftime("%d/%m/%y %H:%M:%S"),
                        calculate_duration(start_time_morning_participant, end_time_morning_participant),
                    ]
                if (row == 1):
                    ws.append([
                        "{} {}".format(participant['nom'].lower(), participant['prenom'].lower()),
                        participant['email'],
                        start_time_afternoon_participant.strftime("%d/%m/%y %H:%M:%S"),
                        end_time_afternoon_participant.strftime("%d/%m/%y %H:%M:%S"),
                        calculate_duration(start_time_afternoon_participant, end_time_afternoon_participant),
                        "Oui",
                        "Oui"
                    ])
                    meetings_and_participants['participants'][participant['email']]['afternoon'] = [
                        "{}".format(participant['nom_complet']),
                        participant['prenom'],
                        participant['email'],
                        start_time_afternoon_participant.strftime("%d/%m/%y %H:%M:%S"),
                        end_time_afternoon_participant.strftime("%d/%m/%y %H:%M:%S"),
                        calculate_duration(start_time_afternoon_participant, end_time_afternoon_participant),
                    ]
                    ws.append(['empty', '', '', '', '', '', 'empty'])
                ws.row_dimensions[ws.max_row].height = 25
        

        virtual_meetings_and_participants['participants'] = meetings_and_participants['participants']

        return virtual_meetings_and_participants

                    
        
    # full_meetings_and_participants['date_debut'] = generate_meetings_and_participants(formation['date_debut'])
    

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    gray_fill = PatternFill(start_color='DCDCDC', end_color='DCDCDC', fill_type='solid')
    bold_font = Font(name='Calibri', size=11, bold=True)
    
    if 'date_debut' in formation:
        date_formation = formation['date_debut']
        sheet_name = date_formation.strftime('%d-%m-%Y')  # Nommer l'onglet
        ws = wb.create_sheet(title=sheet_name)  # Créer un nouvel onglet
        full_meetings_and_participants['date_debut'] = generate_meetings_and_participants(date_formation, ws)
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 'empty' or cell.value:
                    cell.border = thin_border
                    cell.font = Font(name='Helvetica Neue', size=11)
                    if cell.value == 'empty':
                        cell.value = ''
                        if cell.row == ws.max_row:
                            cell.border = None
                
                if cell.value and isinstance(cell.value, str) and '@' in cell.value:
                    # Appliquer le soulignement si la cellule contient un email
                    cell.font = Font(underline='single')

        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value and cell.value != 'N° de réunion' and not cell.value.endswith('zoom'):
                cell.fill = gray_fill
                cell.font = bold_font
   


    # ws.append([])
    # ws.row_dimensions[ws.max_row + 1].height = 80
    # ws.merge_cells(start_row=ws.max_row + 1, start_column=1, end_row=ws.max_row + 1, end_column=7)

    # if (formation.get('date_fin')):
    #    full_meetings_and_participants['date_fin'] = generate_meetings_and_participants(formation['date_fin'], ws.max_row + 1)

    if 'date_fin' in formation:
        date_formation = formation['date_fin']
        sheet_name = date_formation.strftime('%d-%m-%Y')  # Créer un onglet pour la date de fin
        ws = wb.create_sheet(title=sheet_name)
        full_meetings_and_participants['date_fin'] = generate_meetings_and_participants(date_formation, ws)
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 'empty' or cell.value:
                    cell.border = thin_border
                    cell.font = Font(name='Helvetica Neue', size=11)
                    if cell.value == 'empty':
                        cell.value = ''
                
                if cell.value and isinstance(cell.value, str) and '@' in cell.value:
                    # Appliquer le soulignement si la cellule contient un email
                    cell.font = Font(underline='single')
        
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value and cell.value != 'N° de réunion' and not cell.value.endswith('zoom'):
                cell.fill = gray_fill
                cell.font = bold_font


    # Supprimer la feuille par défaut créée par openpyxl
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']


    # Configurer la mise en page pour PDF
    setup_excel_for_pdf(ws)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_directory = os.path.join(script_dir, "../downloads")
    output_file = os.path.join(output_directory, "{}_zoom_timesheet_{}".format(formation['code'], filepath))
    wb.save(output_file)

    return wb, ws, full_meetings_and_participants


