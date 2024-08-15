import os
from openpyxl.styles import Font, PatternFill
import services.source_parser as source_parser
import services.timesheet_generator as timesheet_generator
import services.attendance_certificates_generator as attendance_certificates_generator
import services.virtualclass_synthese_generator as virtualclass_synthese_generator


def generate_timesheet_zoom():
    filepath = 'uploads/'
    full_meetings_and_participants = {}

    for filename in os.listdir(filepath):
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            formation = source_parser.create_formation(filepath + filename)
            participants = source_parser.create_participants(filepath + filename)

            if len(formation) == 0 or len(participants) == 0:
                continue

            wb, ws, full_meetings_and_participants = timesheet_generator.create_zoom_timesheet(filename, formation, participants)
            
            # Appel de la fonction pour générer les tableaux pour chaque demi-journée
            virtualclass_synthese_generator.generate_tables_for_each_meeting(filename, full_meetings_and_participants)

            gray_fill = PatternFill(start_color='DCDCDC', end_color='DCDCDC', fill_type='solid')
            bold_font = Font(name='Calibri', size=11, bold=True)
            for row in range(3, ws.max_row + 1):
                cell = ws.cell(row=row, column=1)
                if cell.value and cell.value != 'N° de réunion' and not cell.value.endswith('zoom'):
                    cell.fill = gray_fill
                    cell.font = bold_font
                    if cell.value == 'empty':
                        cell.value = ''

            wb.save("downloads/{}_zoom_timesheet_{}".format(formation['code'], filename))
            print("zoom_timesheet_{} generated".format(filename))
        else:
            print(filename, 'cannot be used!')
    
    print('Timesheet generated done!')

    return full_meetings_and_participants

def generate_attendance_certificates():
    filepath = 'uploads/'

    for filename in os.listdir(filepath):
        formation = source_parser.create_formation(filepath + filename)
        participants = source_parser.create_participants(filepath + filename)
        if len(formation) == 0 or len(participants) == 0:
            continue

        formation_titre = formation['titre'].replace('\n', '')
        print(f"Formation - {formation_titre}")
        
        for participant in participants:
            attendance_certificates_generator.generate_attendance_certificate(participant, formation)
            print(f"    Attestation de présence généré pour {participant['nom_complet']}")
