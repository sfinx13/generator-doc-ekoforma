import source_parser
import timesheet_generator
import os
from openpyxl.styles import Font, PatternFill

filepath = 'doc/'


for filename in os.listdir(filepath):
    formation = source_parser.create_formation(filepath + filename)
    participants = source_parser.create_participants(filepath + filename)

    wb, ws = timesheet_generator.create_zoom_timesheet(filename, formation, participants)

    gray_fill = PatternFill(start_color='DCDCDC', end_color='DCDCDC', fill_type='solid')
    bold_font = Font(name='Calibri', size=11, bold=True)
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        if cell.value and cell.value != 'N° de réunion' and not cell.value.endswith('zoom'):
            cell.fill = gray_fill
            cell.font = bold_font
            if cell.value == 'empty':
                cell.value = ''

    wb.save("zoom_timesheet_{}".format(filename))
    print("zoom_timesheet_{} generated".format(filename))

print('Done!')